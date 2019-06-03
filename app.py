from flask import Flask, render_template, request, redirect, url_for, jsonify,Markup
from flaskext.mysql import MySQL
import openpyxl
from openpyxl import load_workbook
import plotly.plotly as py
import plotly.tools as tls
import matplotlib.pyplot as plt
from datetime import datetime,timedelta
import time
import calendar
import pandas as pd
from numpy import percentile
from numpy.random import rand

app = Flask(__name__)
mysql = MySQL()
app.config['MYSQL_DATABASE_USER'] = 'tristan'#'peerpower' #
app.config['MYSQL_DATABASE_PASSWORD'] = 'tristan1234'#'pp1234'#
app.config['MYSQL_DATABASE_DB'] = 'tristan' #'pp_benz' #
app.config['MYSQL_DATABASE_HOST'] = 'data-analytics.cnmel21h3vxl.ap-southeast-1.rds.amazonaws.com'#'192.168.9.12'
mysql.init_app(app)

def remainOS():
    rowcount = 1
    column_count = 1
    books = openpyxl.Workbook()
    books.create_sheet('os')
    books.save('os.xlsx')
    sheet = books.get_sheet_by_name('os')
    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date<= curdate() order by loan_app_id asc")
    allID = cursor.fetchall()
    for loan_app_id in allID:
        totallimit = 0
        totaloutstanding = 0
        cursor.execute("select recordable_id,outstanding_balance,loan_limit from loan_records where pmt>0 and ncb_status = 10 and outstanding_balance >0 and recordable_id = %s",loan_app_id[0])
        companyos= cursor.fetchall()
        if cursor.rowcount!=0:

            for row in companyos:
                totaloutstanding += row[1]
                totallimit += row[2]
            ospercent = (totaloutstanding / totallimit) * 100
            sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)
        sheet.cell(row=rowcount, column=1).value = loan_app_id[0]

        rowcount+=1
        cursor.execute("SELECT id FROM shareholders where loan_app_id = %s", loan_app_id)
        allshareholders = cursor.fetchall()
        for rows in allshareholders:
            totallimit = 0
            totaloutstanding = 0
            cursor.execute(
                "select recordable_id,outstanding_balance,loan_limit from loan_records where pmt>0 and ncb_status = 10 and outstanding_balance >0 and recordable_id = %s",
                rows[0])
            companyos = cursor.fetchall()
            if cursor.rowcount != 0:

                for row in companyos:
                    totaloutstanding += row[1]
                    totallimit += row[2]
                ospercent = (totaloutstanding / totallimit) * 100
                sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)

            sheet.cell(row=rowcount, column=1).value = rows[0]
            rowcount += 1
        books.save("os.xlsx");
def bureauyears():
    rowcount = 1
    column_count = 1
    books = openpyxl.Workbook()
    books.create_sheet('os')
    books.save('os.xlsx')
    sheet = books.get_sheet_by_name('os')
    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date<= curdate() order by loan_app_id asc")
    allID = cursor.fetchall()
    for loan_app_id in allID:
        totallimit = 0
        totalmonths = 0
        cursor.execute("select loan_limit,ncb_inquiry_date,date_opened, timestampdiff(MONTH,date_opened,ncb_inquiry_date) as 'months' from loan_records "
                       " inner join company_credit_score_datas on loan_app_id = recordable_id"
                       " where ncb_inquiry_date<curdate() and recordable_id = %s;",loan_app_id[0])
        companyos = cursor.fetchall()
        if cursor.rowcount != 0:

            for row in companyos:
                totalmonths+= row[3]

                totallimit += row[0]
            totalyears = totalmonths / 12.00
            totallimit1 = float(totallimit)
            print(loan_app_id[0])
            print(totallimit1)
            print(totalyears)


            ospercent = ( (totalyears+totallimit1) / totallimit1) * 100.00
            sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)
        sheet.cell(row=rowcount, column=1).value = loan_app_id[0]

        rowcount += 1
        cursor.execute("SELECT id FROM shareholders where loan_app_id = %s", loan_app_id)
        allshareholders = cursor.fetchall()
        for rows in allshareholders:
            totallimit = 0

            cursor.execute(
                "select loan_limit,ncb_inquiry_date,date_opened, timestampdiff(MONTH,date_opened,ncb_inquiry_date) as 'months' from loan_records "
                " inner join company_credit_score_datas on loan_app_id = recordable_id"
                " where ncb_inquiry_date<curdate() and recordable_id = %s;", rows[0])
            companyos = cursor.fetchall()
            if cursor.rowcount != 0:

                for row in companyos:
                    totalmonths += row[3]

                    totallimit += row[0]
                totalyears = totalmonths / 12.00
                totallimit1 = float(totallimit)

                print("shareholder" + rows[0])
                print(totallimit1)
                print(totalyears)
                ospercent = ((totalyears + totallimit1) / totallimit1) * 100.00
                sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)

            sheet.cell(row=rowcount, column=1).value = rows[0]
            rowcount += 1
        books.save("os.xlsx")
def revolving_loans():
    book = openpyxl.Workbook()
    book.create_sheet('Sample')
    book.save('Sample.xlsx')
    conn= mysql.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT distinct(shareholder_id) FROM pp_app_v3_prod.shareholder_credit_score_datas"
                   " inner join loan_records on shareholder_id = recordable_id"
                   " where ncb_inquiry_date <= curdate() and (type_id =2  or type_id = 4) and ncb_status = 10 order by shareholder_id asc")
    shareholders = cursor.fetchall()
    listofshareholder = []
    for manyrows in shareholders:
        listofshareholder.append(manyrows[0])

    rowcount = 1
    for ids in listofshareholder:

        months = 12;
        count = 0
        while months != 0:

            conn = mysql.connect()
            cursor = conn.cursor()
            cursor.execute("select loan_limit,loan_record_payment_histories.outstanding_balance,recordable_id from loan_records"
                           " inner join loan_record_payment_histories on loan_record_id = loan_records.id"
                           " inner join shareholder_credit_score_datas on shareholder_id = recordable_id"
                           " where (loan_records.type_id = 2 or loan_records.type_id = 4) and ncb_status = 10 and recordable_id = %s and date_add(ncb_inquiry_date,INTERVAL -%s MONTH)<=payment_date;",(shareholders[rowcount-1],months))
            data = cursor.fetchall();
            totallimit= 0
            totaloutstanding = 0


            if cursor.rowcount!= 0:

                for rows in data:
                    totallimit = totallimit + rows[0]
                    totaloutstanding = totaloutstanding + rows[1]
                totalopenbal = totallimit - totaloutstanding

                percentage = ((totalopenbal / totallimit) * 100)
                count+= 1;
                sheet = book.get_sheet_by_name('Sample')
                sheet.cell(row=rowcount,column=months).value = str.format('{0:.3f}',percentage)



                print(str(months)+": "+ str.format('{0:.3f}',percentage))


            months -=1;
        sheet.cell(row=rowcount, column=14).value = rows[2]
        rowcount += 1;
        if count == 0:
            print("no records found")
        book.save('Sample.xlsx')
    return 0





    return 0
def dq_percent():
    count = 1
    books = openpyxl.Workbook()
    books.create_sheet('dq')
    books.save('dq.xlsx')
    sheet = books.get_sheet_by_name('dq')
    conn = mysql.connect()
    cursor = conn.cursor()
    cursor.execute("SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date<= curdate() order by loan_app_id asc")
    allID = cursor.fetchall()
    companyid = []
    for id in allID:
        companyid.append(id[0])

    for loan_app_id in companyid:
        dqmonths = [12, 18, 24, 36]
        #this will check for every company first
        #This should check DQ percent for most recent month first and subsequent 12,18,24
        column_count = 2
        for months in dqmonths:
            cursor.execute("select recordable_id,loan_limit,loan_record_payment_histories.outstanding_balance,status_code,ncb_inquiry_date,payment_date from loan_records "
                           " inner join company_credit_score_datas on loan_app_id = recordable_id"
                           " inner join loan_record_payment_histories on loan_record_id = loan_records.id"
                           " where type_id != 4 and recordable_id = %s and loan_record_payment_histories.status_code != 0 and loan_record_payment_histories.status_code != 1 and date_add(ncb_inquiry_date, INTERVAL -%s MONTH) < payment_date",(loan_app_id,months))
            data = cursor.fetchall()
            sheet.cell(row=count, column=1).value = loan_app_id
            totallimit = 0
            totaloutstanding = 0
            for rows in data:

                if cursor.rowcount != 0:
                 #for each row in data retrieved


                    totallimit += rows[1]
                    totaloutstanding += rows[2]
                totalpercentage = (totaloutstanding/totallimit) * 100;
                  # recordable_id
                sheet.cell(row=count, column=column_count).value = str.format('{0:.3f}',totalpercentage)

                books.save('dq.xlsx')
            column_count+=1
        count += 1
        cursor.execute("SELECT id FROM shareholders where loan_app_id = %s", loan_app_id)
        allshareholders = cursor.fetchall()
        column_count = 2
        totallimit = 0
        totaloutstanding = 0
        dqmonths = [12, 18, 24, 36]
        for shareholder in allshareholders:
            sheet.cell(row=count, column=1).value = shareholder[0]
            for shareholdermonths in dqmonths:
                cursor.execute(
                    "select recordable_id,loan_limit,loan_record_payment_histories.outstanding_balance,status_code,ncb_inquiry_date,payment_date from loan_records"
                    " inner join shareholder_credit_score_datas on shareholder_id = recordable_id"
                    " inner join loan_record_payment_histories on loan_record_id = loan_records.id"
                    " where type_id != 4 and recordable_id = %s and loan_record_payment_histories.status_code != 0 and loan_record_payment_histories.status_code != 1 and date_add(ncb_inquiry_date, INTERVAL -%s MONTH) < payment_date",
                    (shareholder[0], shareholdermonths))
                shareholderloans = cursor.fetchall()

                for data in shareholderloans:
                    if cursor.rowcount != 0:
                        totallimit += data[1]
                        totaloutstanding += data[2]
                    stotalpercentage = (totaloutstanding / totallimit) * 100;
                    # recordable_id
                    sheet.cell(row=count, column=column_count).value = str.format('{0:.3f}', stotalpercentage)
                    books.save('dq.xlsx')

                column_count += 1
            count += 1

    books.save('dq.xlsx')
def countyears():
    book = openpyxl.Workbook()

    book.create_sheet('test')
    book.save('test.xlsx')
    sheet = book.get_sheet_by_name('test')
    conn = mysql.connect()
    cursor = conn.cursor()
    rowcount = 1
    cursor.execute("SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date<= curdate() order by loan_app_id asc")
    companyids = cursor.fetchall()

    for company_id in companyids:
        #company query
        cursor.execute("select recordable_id,loan_limit,ncb_inquiry_date,date_opened, timestampdiff(MONTH,date_opened,ncb_inquiry_date) as 'months' from loan_records "
                       " inner join company_credit_score_datas on loan_app_id = recordable_id "
                       "where ncb_inquiry_date<curdate() and recordable_id = %s and loan_limit>0;",company_id[0])
        data = cursor.fetchall()
        totalmonths = 0.0
        totalyear = 0.0
        loanvariable = 0.0
        totallimit = 0.0
        if cursor.rowcount!= 0:
            for rows in data:

                loanvariable += float((float(rows[1]) * float(rows[4]/12)))
                totallimit += float(rows[1])
                print(rows[0])
                print(loanvariable)
                print(totallimit)
            ospercent = (loanvariable / totallimit) *100
            sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)
        sheet.cell(row=rowcount, column=1).value = company_id[0]
        rowcount+=1

        cursor.execute("select id from shareholders where loan_app_id = %s",company_id[0])
        shareholderids = cursor.fetchall()
        for shareholder_id in shareholderids:
            cursor.execute(
                "select recordable_id,loan_limit,ncb_inquiry_date,date_opened, timestampdiff(MONTH,date_opened,ncb_inquiry_date) as 'months' from loan_records "
                " inner join shareholder_credit_score_datas on shareholder_id = recordable_id "
                "where ncb_inquiry_date<curdate() and recordable_id = %s and loan_limit>0;", shareholder_id[0])
            data = cursor.fetchall()

            loanvariable = 0.0
            totallimit = 0.0
            if cursor.rowcount != 0:
                for rows in data:
                    loanvariable += float((float(rows[1]) * float(rows[4]/12)))
                    totallimit += float(rows[1])
                ospercent = (loanvariable / totallimit) * 100
                sheet.cell(row=rowcount, column=2).value = str.format('{0:.3f}', ospercent)
            sheet.cell(row=rowcount, column=1).value = shareholder_id[0];
            rowcount+=1
    book.save("test.xlsx");
def loan_purpose():
    book = openpyxl.Workbook()
    book.create_sheet('purpose')
    book.save('purpose.xlsx')
    sheet = book.get_sheet_by_name('purpose')
    conn = mysql.connect()
    cursor = conn.cursor()
    rowcount = 1
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date<= curdate() order by loan_app_id asc")
    companyids = cursor.fetchall()

    for company_id in companyids:
        # company query
        cursor.execute("SELECT loan_apps.id,loan_purposes.desc_en FROM pp_app_v3_prod.loan_apps"
                       " inner join loan_purposes on loan_purpose_id = loan_purposes.id"
                       " where loan_apps.id = %s",company_id[0])
        data = cursor.fetchall()
        if cursor.rowcount != 0:
            for rows in data:
                sheet.cell(row=rowcount, column=2).value = rows[1]
        sheet.cell(row=rowcount, column=1).value = company_id[0]
        rowcount += 1

        cursor.execute("select id from shareholders where loan_app_id = %s", company_id[0])
        shareholderids = cursor.fetchall()
        for shareholder_id in shareholderids:
            sheet.cell(row=rowcount, column=1).value = shareholder_id[0];
            rowcount += 1
    book.save("purpose.xlsx");
def approval():
    book = openpyxl.Workbook()
    book.create_sheet('goodbad')
    book.save('gooadbad.xlsx')
    sheet = book.get_sheet_by_name('goodbad')
    conn = mysql.connect()
    cursor = conn.cursor()
    rowcount = 1
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date <= curdate() and (loan_app_id != 44009759) and (loan_app_id != 44009564) order by loan_app_id asc")
    companyids = cursor.fetchall()

    for company_id in companyids:
        # company query
        cursor.execute("select id from investment_orders where loan_contract_id = %s;",company_id[0])
        data = cursor.fetchall()
        if cursor.rowcount != 0:
            companytrust = True
            for rows in data:
                cursor.execute("SELECT DATE(due_date),date(paid_at),timestampdiff(day,due_date,paid_at) as 'months' FROM pp_app_v3_prod.investor_amortization_schedules "
                               "where investment_order_id = %s and CAST(due_date as DATE) < CAST(paid_at as DATE) ;",rows[0])
                companylate = cursor.fetchall()
                if cursor.rowcount < 2:
                    companytrust = True
                else:
                    companytrust= False
            if companytrust == True:
                sheet.cell(row=rowcount, column=2).value = "Good"
        sheet.cell(row=rowcount, column=1).value = company_id[0]
        rowcount += 1

        cursor.execute("select id from shareholders where loan_app_id = %s", company_id[0])
        shareholderids = cursor.fetchall()
        for shareholder_id in shareholderids:
            sheet.cell(row=rowcount, column=1).value = shareholder_id[0];
            rowcount += 1
    book.save("goodbad.xlsx");
def goodbad():
    book = openpyxl.Workbook()
    book.create_sheet('approval')
    book.save('approval.xlsx')
    sheet = book.get_sheet_by_name('approval')
    conn = mysql.connect()
    cursor = conn.cursor()
    rowcount = 1
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date <= curdate() and (loan_app_id != 44009759) and (loan_app_id != 44009564) order by loan_app_id asc")
    companyids = cursor.fetchall()

    for company_id in companyids:
        # company query
        cursor.execute("SELECT loan_app_id FROM pp_app_v3_prod.loan_contracts where loan_app_id = %s;", company_id[0])
        data = cursor.fetchall()
        if cursor.rowcount != 0:
            for rows in data:
                sheet.cell(row=rowcount, column=2).value = "Approved"
                shareholder_approved = True;
        else:
            sheet.cell(row=rowcount, column=2).value = "Rejected"
            shareholder_approved = False;
        sheet.cell(row=rowcount, column=1).value = company_id[0]
        rowcount += 1

        cursor.execute("select id from shareholders where loan_app_id = %s", company_id[0])
        shareholderids = cursor.fetchall()
        for shareholder_id in shareholderids:
            sheet.cell(row=rowcount, column=1).value = shareholder_id[0];
            if shareholder_approved == True:
                sheet.cell(row=rowcount, column=2).value = "Approved"
            else:
                sheet.cell(row=rowcount, column=2).value = "Rejected"
            rowcount += 1
    book.save("approval.xlsx");
def company_score():
    book = openpyxl.Workbook()
    book.create_sheet('score')
    book.save('score.xlsx')
    sheet = book.get_sheet_by_name('score')
    conn = mysql.connect()
    cursor = conn.cursor()
    rowcount = 1
    cursor.execute(
        "SELECT loan_app_id FROM pp_app_v3_prod.company_credit_score_datas where ncb_inquiry_date <= curdate() and (loan_app_id != 44009759) and (loan_app_id != 44009564) order by loan_app_id asc")
    companyids = cursor.fetchall()

    for company_id in companyids:
        # company query
        cursor.execute("select credit_score_group from loan_apps where id = %s;", company_id[0])
        data = cursor.fetchall()
        if cursor.rowcount != 0:
            for rows in data:
                sheet.cell(row=rowcount, column=2).value = rows[0]
        sheet.cell(row=rowcount, column=1).value = company_id[0]
        rowcount += 1

        cursor.execute("select id from shareholders where loan_app_id = %s", company_id[0])
        shareholderids = cursor.fetchall()
        for shareholder_id in shareholderids:
            sheet.cell(row=rowcount, column=1).value = shareholder_id[0];
            rowcount += 1
    book.save("score.xlsx");



weeklabels = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
hourlabels = ['6:00','7:00','8::00','09:00','10:00','11:00','12:00','13:00','14:00','15:00','16:00','17:00','18:00','19:00','20:00','21:00','22:00','23:00','00:00']
monthlabels= ["January 18'","February 18'","March 18'","April 18'","May 18'","June 18'","July 18'","August 18'","September 18'","October 18'","November 18'","December 18'","January 19'","February 19'","March 19'"]
agentlabels = ["101","102","103","104","105","106","107","108","109"]
totalvalues = []
answered = []
notanswered = []
busy = []
failed =[]
inbound=[]
outbound=[]
totalinbound = 0
totaloutbound = 0
totalcalls = 0
successvalue = []

month1 =[]
month2 =[]
month3 =[]
month4 =[]
month5 =[]
month6 =[]
month7 =[]
month8 =[]
month9 =[]
month10 =[]
month11 =[]
month12 =[]
month13 =[]
month14 =[]
month15 =[]
biglist = [month1,month2,month3,month4,month5,month6,month7,month8,month9,month10,month11,month12,month13,month14,month15]
answered101= []
noanswer101 = []
busy101 = []
failed101 =[]
answered102= []
noanswer102 = []
busy102 = []
failed102 =[]
answered103= []
noanswer103 = []
busy103 = []
failed103 =[]
answered104= []
noanswer104 = []
busy104 = []
failed104 =[]
answered105= []
noanswer105 = []
busy105 = []
failed105 =[]
answered106= []
noanswer106 = []
busy106 = []
failed106 =[]
answered107= []
noanswer107 = []
busy107 = []
failed107 =[]
answered108= []
noanswer108 = []
busy108 = []
failed108 =[]
answered109= []
noanswer109 = []
busy109 = []
failed109 =[]
agent101=[answered101,noanswer101,busy101,failed101]
agent102=[answered102,noanswer102,busy102,failed102]
agent103=[answered103,noanswer103,busy103,failed103]
agent104=[answered104,noanswer104,busy104,failed104]
agent105=[answered105,noanswer105,busy105,failed105]
agent106=[answered106,noanswer106,busy106,failed106]
agent107=[answered107,noanswer107,busy107,failed107]
agent108=[answered108,noanswer108,busy108,failed108]
agent109=[answered109,noanswer109,busy109,failed109]

agentslist = [agent101,agent102,agent103,agent104,agent105,agent106,agent107,agent108,agent109]
testlist = []
testlist1 =[]
def agentsmonths(startdate,enddate):
    data = pd.read_csv("cdr.csv", low_memory=False, encoding='latin-1')
    answered101.clear();
    answered102.clear();
    answered103.clear();
    answered104.clear();
    answered105.clear();
    answered106.clear();
    answered107.clear();
    answered108.clear();
    answered109.clear();
    noanswer101.clear();
    noanswer102.clear();
    noanswer103.clear();
    noanswer104.clear();
    noanswer105.clear();
    noanswer106.clear();
    noanswer107.clear();
    noanswer108.clear();
    noanswer109.clear();
    busy101.clear();
    busy102.clear();
    busy103.clear();
    busy104.clear();
    busy105.clear();
    busy106.clear();
    busy107.clear();
    busy108.clear();
    busy109.clear();
    failed101.clear();
    failed102.clear();
    failed103.clear();
    failed104.clear();
    failed105.clear();
    failed106.clear();
    failed107.clear();
    failed108.clear();
    failed109.clear();

    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    monthlabels.clear()
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    haha = pd.to_datetime(data['Date'])
    m1 = haha.map(lambda x: 100*x.year + x.month)
    m1=m1.unique()
    m1[:] = m1[::-1]
    for month in m1:
        x = str(month)
        monthlabels.append(x[:4]+'-'+x[4:])
    date = pd.to_datetime((data["Date"]))

    for months in monthlabels:
        y = months.split('-')
        yearcount = int(y[0])
        monthcount = int(y[1])
        for agents,numbers in zip(agentslist,agentlabels):
            isoutbound = (data.Source == numbers) & (data.Destination != '101') & (data.Destination != '102') & (
                    data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                                 data.Destination != '106') & (data.Destination != '107') & (
                                 data.Destination != '108') & (
                                 data.Destination != '109')
            isinbound = (data.Destination == numbers) & (data.Source != '101') & (data.Source != '102') & (
                    data.Source != '103') & (
                                data.Source != '104') & (
                                data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                                data.Source != '108') & (
                                data.Source != '109')
            agents[0].append(len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))
            agents[1].append(len(data[
                            (data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                                isoutbound)]) + len(data[(data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))
            agents[2].append(len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(
                        data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isinbound)]))
            agents[3].append(len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))
    print(agentslist)
def agentweeks(startdate,enddate):
    data = pd.read_csv("cdr.csv", low_memory=False, encoding='latin-1')
    answered101.clear();
    answered102.clear();
    answered103.clear();
    answered104.clear();
    answered105.clear();
    answered106.clear();
    answered107.clear();
    answered108.clear();
    answered109.clear();
    noanswer101.clear();
    noanswer102.clear();
    noanswer103.clear();
    noanswer104.clear();
    noanswer105.clear();
    noanswer106.clear();
    noanswer107.clear();
    noanswer108.clear();
    noanswer109.clear();
    busy101.clear();
    busy102.clear();
    busy103.clear();
    busy104.clear();
    busy105.clear();
    busy106.clear();
    busy107.clear();
    busy108.clear();
    busy109.clear();
    failed101.clear();
    failed102.clear();
    failed103.clear();
    failed104.clear();
    failed105.clear();
    failed106.clear();
    failed107.clear();
    failed108.clear();
    failed109.clear();

    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    monthlabels.clear()
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    haha = pd.to_datetime(data['Date'])
    m1 = haha.map(lambda x: 100 * x.year + x.month)
    m1 = m1.unique()
    m1[:] = m1[::-1]
    for month in m1:
        x = str(month)
        monthlabels.append(x[:4] + '-' + x[4:])
    date = pd.to_datetime((data["Date"]))

    for months in monthlabels:
        y = months.split('-')
        yearcount = int(y[0])
        monthcount = int(y[1])
        for agents, numbers in zip(agentslist, agentlabels):
            isoutbound = (data.Source == numbers) & (data.Destination != '101') & (data.Destination != '102') & (
                    data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                                 data.Destination != '106') & (data.Destination != '107') & (
                                 data.Destination != '108') & (
                                 data.Destination != '109')
            isinbound = (data.Destination == numbers) & (data.Source != '101') & (data.Source != '102') & (
                    data.Source != '103') & (
                                data.Source != '104') & (
                                data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                                data.Source != '108') & (
                                data.Source != '109')
            agents[0].append(
                len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                    isoutbound)]) + len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (
                        date.dt.year == yearcount) & (isinbound)]))
            agents[1].append(len(data[
                                     (data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (
                                                 date.dt.year == yearcount) & (
                                         isoutbound)]) + len(
                data[(data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (
                        date.dt.year == yearcount) & (isinbound)]))
            agents[2].append(
                len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                    isoutbound)]) + len(
                    data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isinbound)]))
            agents[3].append(
                len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                    isoutbound)]) + len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (
                        date.dt.year == yearcount) & (isinbound)]))
    print(agentslist)
def writestatus():
    data = pd.read_csv("C:\\Users\\PP-Tristan\\Documents\\SIMPLE\\cdrmay.csv",low_memory=False,encoding='latin-1')
    date = pd.to_datetime((data["Date"]))
    answeredtoday = data[((data.Call_Status == 'Answered')& (data.Dates == '5/9/2019'))]
    missedtoday = data[((data.Call_Status == 'Not Answered') & (data.Dates == '5/9/2019'))]
    print(len(answeredtoday))
    print(len(missedtoday))
def alldataboxplot(status):
    totalduration = 0
    data = pd.read_csv("cdr.csv",low_memory=False,encoding='latin-1')

    isoutbound = ((data.Source == '101') | (data.Source == '102') | (data.Source == '103') | (data.Source == '104') | (
                data.Source == '105') | (data.Source == '106') | (data.Source == '107') | (data.Source == '108') | (
                              data.Source == '109') & (data.Destination != '101')& (data.Destination != '102')& (data.Destination != '103')& (data.Destination != '104')& (data.Destination != '105')& (data.Destination != '106')& (data.Destination != '107')& (data.Destination != '108')& (data.Destination != '109'))
    isinbound = ((data.Source != '101') & (data.Source != '102') & (data.Source != '103') & (data.Source != '104') & (
                data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (data.Source != '108') & (
                             data.Source != '109') & (data.Destination != '9001')& (data.Destination != '9002')& (data.Destination != '9003')& (data.Destination != '9004') & (data.Destination != 's')& (data.Destination != 'hangup'))
    answeredcalls = data[(data.Status == 'ANSWERED')]
    date = pd.to_datetime((data["Date"]))
    count=1
    secondcount = 1
    if status =='inbound':
        for monthlist in biglist:
            monthlist.clear()
            if count <= 12:
                dura = (
                data["Duration"][(data.Status == 'ANSWERED') & (date.dt.month == count) & (date.dt.year == 2018)& (isinbound)])
                duration = dura.str.split('s', 1)
                x = (duration.apply(lambda x: x[0])).astype(int)
                biglist[count - 1] = x.tolist()
                biglist[count - 1].sort()

                if len(biglist[count - 1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])
                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count - 1] = [minimum, quartiles[0], quartiles[1], quartiles[2], maximum]

                count += 1

            else:
                dura = (
                data["Duration"][(data.Status == 'ANSWERED') & (date.dt.month == secondcount) & (date.dt.year == 2019)& (isoutbound)])
                duration = dura.str.split('s', 1)
                x = (duration.apply(lambda x: x[0])).astype(int)
                biglist[count - 1] = x.tolist()
                biglist[count - 1].sort()

                if len(biglist[count - 1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])

                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count - 1] = [minimum, quartiles[0], quartiles[1], quartiles[2], maximum]
                secondcount += 1
                count += 1

    elif status =='all':
        for monthlist in biglist:
            if count <=12:
                dura = (data["Duration"][(data.Status == 'ANSWERED')& (date.dt.month==count) & (date.dt.year == 2018)])
                duration = dura.str.split('s',1)
                x= (duration.apply(lambda x: x[0])).astype(int)
                biglist[count-1] = x.tolist()
                biglist[count-1].sort()

                if len(biglist[count - 1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])
                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count - 1] = [minimum, quartiles[0], quartiles[1], quartiles[2], maximum]


                count+=1

            else:
                dura = (data["Duration"][(data.Status == 'ANSWERED') & (date.dt.month == secondcount) & (date.dt.year == 2019)])
                duration = dura.str.split('s', 1)
                x = (duration.apply(lambda x: x[0])).astype(int)
                biglist[count-1] = x.tolist()
                biglist[count-1].sort()

                if len(biglist[count-1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])

                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count-1] = [minimum,quartiles[0],quartiles[1],quartiles[2],maximum]
                secondcount += 1
                count+=1
    else:
        for monthlist in biglist:
            if count <= 12:
                dura = (
                data["Duration"][(data.Status == 'ANSWERED') & (date.dt.month == count) & (date.dt.year == 2018)& isoutbound])
                duration = dura.str.split('s', 1)
                x = (duration.apply(lambda x: x[0])).astype(int)
                biglist[count - 1] = x.tolist()
                biglist[count - 1].sort()

                if len(biglist[count - 1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])
                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count - 1] = [minimum, quartiles[0], quartiles[1], quartiles[2], maximum]

                count += 1

            else:
                dura = (
                data["Duration"][(data.Status == 'ANSWERED') & (date.dt.month == secondcount) & (date.dt.year == 2019)& isoutbound])
                duration = dura.str.split('s', 1)
                x = (duration.apply(lambda x: x[0])).astype(int)
                biglist[count - 1] = x.tolist()
                biglist[count - 1].sort()

                if len(biglist[count - 1]) > 0:
                    minimum = min(biglist[count - 1])
                    maximum = max(biglist[count - 1])

                    quartiles = percentile(biglist[count - 1], [25, 50, 75])
                    biglist[count - 1] = [minimum, quartiles[0], quartiles[1], quartiles[2], maximum]
                secondcount += 1
                count += 1
def agentdata(status,startdate,enddate):
    data = pd.read_csv("cdr.csv", low_memory=False, encoding='latin-1')
    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    isoutbound = ((data.Source == '101') | (data.Source == '102') | (data.Source == '103') | (
            data.Source == '104') | (
                          data.Source == '105') | (data.Source == '106') | (data.Source == '107') | (
                          data.Source == '108') | (
                          data.Source == '109') & (data.Destination != '101') & (data.Destination != '102') & (
                          data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                          data.Destination != '106') & (data.Destination != '107') & (data.Destination != '108') & (
                          data.Destination != '109'))
    isinbound = ((data.Source != '101') & (data.Source != '102') & (data.Source != '103') & (
            data.Source != '104') & (
                         data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                         data.Source != '108') & (
                         data.Source != '109') & (data.Destination != '9001') & (data.Destination != '9002') & (
                         data.Destination != '9003') & (data.Destination != '9004') & (data.Destination != 's') & (
                         data.Destination != 'hangup'))
    if status == 'inbound':
        for agent in agentlabels:
            answered.append(len(data[(data.Status == "ANSWERED")& (data.Destination == agent)& isinbound]))
            notanswered.append(len(data[(data.Status == "NO ANSWER")& (data.Destination == agent)&isinbound]))
            busy.append(len(data[(data.Status == "BUSY")& (data.Destination == agent)&isinbound]))
            failed.append(len(data[(data.Status == "FAILED")& (data.Destination == agent)&isinbound]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);


    elif status == 'outbound':
        for agent in agentlabels:
            answered.append(len(data[(data.Status == "ANSWERED") & (data.Source == agent) & isoutbound]))
            notanswered.append(len(data[(data.Status == "NO ANSWER") & (data.Source == agent) & isoutbound]))
            busy.append(len(data[(data.Status == "BUSY") & (data.Source == agent) & isoutbound]))
            failed.append(len(data[(data.Status == "FAILED") & (data.Source == agent) & isoutbound]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isoutbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isoutbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isoutbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isoutbound)]);
    else:
        for agent in agentlabels:
            answered.append(len(data[(data.Status == "ANSWERED") & (data.Destination == agent)&isinbound])+len(data[(data.Status == "ANSWERED") & (data.Source == agent) & isoutbound]))
            notanswered.append(len(data[(data.Status == "NO ANSWER") & (data.Destination == agent)&isinbound])+len(data[(data.Status == "NO ANSWER") & (data.Source == agent) & isoutbound]))
            busy.append(len(data[(data.Status == "BUSY") & (data.Destination == agent)&isinbound ])+len(data[(data.Status == "BUSY") & (data.Source == agent) & isoutbound]))
            failed.append(len(data[(data.Status == "FAILED") & (data.Destination == agent)&isinbound ])+len(data[(data.Status == "FAILED") & (data.Source == agent) & isoutbound]))
            totalanswered = len(
                data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                data[(data.Status == 'ANSWERED') & isinbound])

            totalnoanswer = len(
                data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                data[(data.Status == 'NO ANSWER') & isinbound])
            totalbusy = len(
                data[(data.Status == 'BUSY') & isoutbound]) + len(
                data[(data.Status == 'BUSY') & isinbound])
            totalfailed = len(
                data[(data.Status == 'FAILED') & isoutbound]) + len(
                data[(data.Status == 'FAILED') & isinbound])
    totalvalues.append(totalanswered)
    totalvalues.append(totalnoanswer)
    totalvalues.append(totalbusy)
    totalvalues.append(totalfailed)
    if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
        totalvalues.append(0)
    else:
        totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))
def alldatamonths(agent,status,startdate,enddate):
    data = pd.read_csv("cdr.csv", low_memory=False, encoding='latin-1')
    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    monthlabels.clear()
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    haha = pd.to_datetime(data['Date'])
    m1 = haha.map(lambda x: 100*x.year + x.month)
    m1=m1.unique()
    m1[:] = m1[::-1]
    for month in m1:
        x = str(month)
        monthlabels.append(x[:4]+'-'+x[4:])
    date = pd.to_datetime((data["Date"]))
    if agent == 'all':
        isoutbound = ((data.Source == '101') | (data.Source == '102') | (data.Source == '103') | (
                    data.Source == '104') | (
                              data.Source == '105') | (data.Source == '106') | (data.Source == '107') | (
                                  data.Source == '108') | (
                              data.Source == '109') & (data.Destination != '101') & (data.Destination != '102') & (
                              data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                              data.Destination != '106') & (data.Destination != '107') & (data.Destination != '108') & (
                              data.Destination != '109'))
        isinbound = ((data.Source != '101') & (data.Source != '102') & (data.Source != '103') & (
                    data.Source != '104') & (
                             data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                                 data.Source != '108') & (
                             data.Source != '109') & (data.Destination != '9001') & (data.Destination != '9002') & (
                             data.Destination != '9003') & (data.Destination != '9004') & (data.Destination != 's') & (
                             data.Destination != 'hangup'))

        count = 1
        secondcount=1
        if status == 'inbound':
            for month in monthlabels:
                y=month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == y[1])&(date.dt.year ==y[0]) & (isinbound)]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.month == y[1])&(date.dt.year ==y[0]) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.month == y[1])&(date.dt.year ==y[0]) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == y[1])&(date.dt.year ==y[0]) & (isinbound)]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);

        elif status == 'outbound':
            for month in monthlabels:
                y = month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == y[1]) & (date.dt.year == y[0]) & (isoutbound)]))
                notanswered.append(
                    len(data[
                            (data.Status == 'NO ANSWER') & (date.dt.month == y[1]) & (date.dt.year == y[0]) & (isoutbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.month == y[1]) & (date.dt.year == y[0]) & (isoutbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == y[1]) & (date.dt.year == y[0]) & (isoutbound)]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isoutbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isoutbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isoutbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isoutbound)]);

        else:
            for month in monthlabels:
                y = month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])


                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isoutbound)])+len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isinbound)]))

                notanswered.append(
                    len(data[
                            (data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isoutbound)])+len(data[(data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isoutbound)])+len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isoutbound)])+len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year ==yearcount) & (isinbound)]))
            totalanswered = len(
                data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                data[(data.Status == 'ANSWERED') & isinbound])

            totalnoanswer = len(
                data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                data[(data.Status == 'NO ANSWER') & isinbound])
            totalbusy = len(
                data[(data.Status == 'BUSY') & isoutbound]) + len(
                data[(data.Status == 'BUSY') & isinbound])
            totalfailed = len(
                data[(data.Status == 'FAILED') & isoutbound]) + len(
                data[(data.Status == 'FAILED') & isinbound])
        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))

    else:
        isoutbound = (data.Source == agent) & (data.Destination != '101') & (data.Destination != '102') & (
                data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                             data.Destination != '106') & (data.Destination != '107') & (
                             data.Destination != '108') & (
                             data.Destination != '109')
        isinbound = (data.Destination == agent) & (data.Source != '101') & (data.Source != '102') & (
                data.Source != '103') & (
                            data.Source != '104') & (
                            data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                            data.Source != '108') & (
                            data.Source != '109')
        date = pd.to_datetime((data["Date"]))

        if status == 'inbound':
            for month in monthlabels:
                y = month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isinbound)]))
                print(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isinbound)])
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isinbound)]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);

        elif status == 'outbound':
            for month in monthlabels:
                y = month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]))
                notanswered.append(
                    len(data[
                            (data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                                isoutbound)]))
                busy.append(
                    len(data[
                            (data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isoutbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]))
            totalanswered = len(data[(data.Status == 'ANSWERED') & (isoutbound)]);
            totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isoutbound)]);
            totalbusy = len(data[(data.Status == 'BUSY') & (isoutbound)]);
            totalfailed = len(data[(data.Status == 'FAILED') & (isoutbound)]);

        else:
            for month in monthlabels:
                y = month.split('-')
                yearcount = int(y[0])
                monthcount = int(y[1])
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(data[(data.Status == 'ANSWERED') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))

                notanswered.append(
                    len(data[
                            (data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                                isoutbound)]) + len(data[(data.Status == 'NO ANSWER') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(
                        data[(data.Status == 'BUSY') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (date.dt.year == yearcount) & (
                        isoutbound)]) + len(data[(data.Status == 'FAILED') & (date.dt.month == monthcount) & (
                                date.dt.year == yearcount) & (isinbound)]))
            totalanswered = len(
                data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                data[(data.Status == 'ANSWERED') & isinbound])

            totalnoanswer = len(
                data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                data[(data.Status == 'NO ANSWER') & isinbound])
            totalbusy = len(
                data[(data.Status == 'BUSY') & isoutbound]) + len(
                data[(data.Status == 'BUSY') & isinbound])
            totalfailed = len(
                data[(data.Status == 'FAILED') & isoutbound]) + len(
                data[(data.Status == 'FAILED') & isinbound])

        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))
def alldatahours(agent,status,startdate,enddate):
    data = pd.read_csv("cdr.csv",low_memory=False,encoding='latin-1')
    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    if agent == 'all':
        isoutbound = ((data.Source == '101') | (data.Source == '102') | (data.Source == '103') | (data.Source == '104') | (
                data.Source == '105') | (data.Source == '106') | (data.Source == '107') | (data.Source == '108') | (
                              data.Source == '109') & (data.Destination != '101') & (data.Destination != '102') & (
                                  data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                                  data.Destination != '106') & (data.Destination != '107') & (data.Destination != '108') & (
                                  data.Destination != '109'))
        isinbound = ((data.Source != '101') & (data.Source != '102') & (data.Source != '103') & (data.Source != '104') & (
                data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (data.Source != '108') & (
                             data.Source != '109') & (data.Destination != '9001') & (data.Destination != '9002') & (
                                 data.Destination != '9003') & (data.Destination != '9004') & (data.Destination != 's') & (
                                 data.Destination != 'hangup'))
        date = pd.to_datetime((data["Date"]))

        count = 6
        if status == 'inbound':
            for hour in hourlabels:
                if count == 24:
                    count = 0

                answeredcall = len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & (isinbound)])
                allcalls = (len(data[(date.dt.hour == count) & (isinbound)]))

                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & (isinbound)]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.hour == count) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.hour == count) & (isinbound)]))
                totalanswered = len(data[(data.Status == 'ANSWERED')& (isinbound)]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER')& (isinbound)]);
                totalbusy = len(data[(data.Status == 'BUSY')& (isinbound)]);
                totalfailed = len(data[(data.Status == 'FAILED')& (isinbound)]);
                count += 1

        elif status == 'outbound':
            for hour in hourlabels:
                if count == 24:
                    count = 0
                answeredcall = len(
                    data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound])
                allcalls = (len(data[(date.dt.hour == count) & isoutbound]))
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isoutbound]))
                busy.append(len(data[(data.Status == 'BUSY') & (date.dt.hour == count) & isoutbound]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.hour == count) & isoutbound]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & isoutbound]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & isoutbound]);
                totalbusy = len(data[(data.Status == 'BUSY') & isoutbound]);
                totalfailed = len(data[(data.Status == 'FAILED') & isoutbound]);
                count += 1

        else:
            for hour in hourlabels:
                if count == 24:
                    count = 0
                actualanswered = len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isinbound])
                answered.append(actualanswered)
                actualnotanswered = len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isinbound])
                notanswered.append(actualnotanswered)
                actualbusy = len(
                    data[(data.Status == 'BUSY') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & (date.dt.hour == count) & isinbound])
                busy.append(actualbusy)
                actualfailed = len(
                    data[(data.Status == 'FAILED') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & (date.dt.hour == count) & isinbound])
                failed.append(actualfailed)

                totalanswered = len(
                    data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & isinbound])

                totalnoanswer = len(
                    data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & isinbound])
                totalbusy = len(
                    data[(data.Status == 'BUSY') & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & isinbound])
                totalfailed = len(
                    data[(data.Status == 'FAILED') & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & isinbound])
                count += 1

        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))
    else:
        isoutbound = (data.Source == agent ) & (data.Destination != '101') & (data.Destination != '102') & (
                data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                             data.Destination != '106') & (data.Destination != '107') & (
                             data.Destination != '108') & (
                             data.Destination != '109')
        isinbound = (data.Destination == agent ) & (data.Source != '101') & (data.Source != '102') & (
                    data.Source != '103') & (
                            data.Source != '104') & (
                            data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                                data.Source != '108') & (
                            data.Source != '109')
        date = pd.to_datetime((data["Date"]))

        count = 6
        if status == 'inbound':
            for hour in hourlabels:
                if count == 24:
                    count = 0

                answeredcall = len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & (isinbound)])
                allcalls = (len(data[(date.dt.hour == count) & (isinbound)]))

                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & (isinbound)]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.hour == count) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.hour == count) & (isinbound)]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
                totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
                totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);
                count += 1

        elif status == 'outbound':
            for hour in hourlabels:
                if count == 24:
                    count = 0
                answeredcall = len(
                    data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound])
                allcalls = (len(data[(date.dt.hour == count) & isoutbound]))
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isoutbound]))
                busy.append(len(data[(data.Status == 'BUSY') & (date.dt.hour == count) & isoutbound]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.hour == count) & isoutbound]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & isoutbound]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & isoutbound]);
                totalbusy = len(data[(data.Status == 'BUSY') & isoutbound]);
                totalfailed = len(data[(data.Status == 'FAILED') & isoutbound]);
                count += 1

        else:
            for hour in hourlabels:
                if count == 24:
                    count = 0
                actualanswered = len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isoutbound])+ len(data[(data.Status == 'ANSWERED') & (date.dt.hour == count) & isinbound])
                answered.append(actualanswered)
                actualnotanswered = len(data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.hour == count) & isinbound])
                notanswered.append(actualnotanswered)
                actualbusy = len(
                    data[(data.Status == 'BUSY') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & (date.dt.hour == count) & isinbound])
                busy.append(actualbusy)
                actualfailed = len(
                    data[(data.Status == 'FAILED') & (date.dt.hour == count) & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & (date.dt.hour == count) & isinbound])
                failed.append(actualfailed)

                totalanswered = len(
                    data[(data.Status == 'ANSWERED')& isoutbound]) + len(
                    data[(data.Status == 'ANSWERED')& isinbound])

                totalnoanswer = len(
                    data[(data.Status == 'NO ANSWER')& isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER')& isinbound])
                totalbusy = len(
                    data[(data.Status == 'BUSY')& isoutbound]) + len(
                    data[(data.Status == 'BUSY')& isinbound])
                totalfailed = len(
                    data[(data.Status == 'FAILED')& isoutbound]) + len(
                    data[(data.Status == 'FAILED')& isinbound])
                count += 1

        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))
def alldatadays(agent,status,startdate,enddate):
    data = pd.read_csv("cdr.csv", low_memory=False, encoding='latin-1')
    totalvalues.clear();
    answered.clear();
    notanswered.clear();
    busy.clear();
    failed.clear();
    successvalue.clear();
    data = data[(data.Date >= startdate) & (data.Date <= enddate)]
    if agent == 'all':
        isoutbound = ((data.Source == '101') | (data.Source == '102') | (data.Source == '103') | (
                    data.Source == '104') | (
                              data.Source == '105') | (data.Source == '106') | (data.Source == '107') | (
                                  data.Source == '108') | (
                              data.Source == '109') & (data.Destination != '101') & (data.Destination != '102') & (
                              data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                              data.Destination != '106') & (data.Destination != '107') & (data.Destination != '108') & (
                              data.Destination != '109'))
        isinbound = ((data.Source != '101') & (data.Source != '102') & (data.Source != '103') & (
                    data.Source != '104') & (
                             data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                                 data.Source != '108') & (
                             data.Source != '109') & (data.Destination != '9001') & (data.Destination != '9002') & (
                             data.Destination != '9003') & (data.Destination != '9004') & (data.Destination != 's') & (
                             data.Destination != 'hangup'))
        date = pd.to_datetime((data["Date"]))

        count = 0
        if status == 'inbound':
            for day in weeklabels:
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & (isinbound)]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.weekday == count) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.weekday == count) & (isinbound)]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
                totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
                totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);
                count += 1

        elif status == 'outbound':
            for day in weeklabels:
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isoutbound]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isoutbound]))
                busy.append(len(data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isoutbound]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isoutbound]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & isoutbound]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & isoutbound]);
                totalbusy = len(data[(data.Status == 'BUSY') & isoutbound]);
                totalfailed = len(data[(data.Status == 'FAILED') & isoutbound]);
                count += 1

        else:
            for day in weeklabels:

                actualanswered = len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isinbound])

                answered.append(actualanswered)
                actualnotanswered = len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isinbound])
                notanswered.append(actualnotanswered)
                actualbusy = len(
                    data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isinbound])
                busy.append(actualbusy)
                actualfailed = len(
                    data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isinbound])
                failed.append(actualfailed)

                totalanswered = len(
                    data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & isinbound])

                totalnoanswer = len(
                    data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & isinbound])
                totalbusy = len(
                    data[(data.Status == 'BUSY') & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & isinbound])
                totalfailed = len(
                    data[(data.Status == 'FAILED') & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & isinbound])
                count += 1

        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))
    else:
        isoutbound = (data.Source == agent) & (data.Destination != '101') & (data.Destination != '102') & (
                data.Destination != '103') & (data.Destination != '104') & (data.Destination != '105') & (
                             data.Destination != '106') & (data.Destination != '107') & (
                             data.Destination != '108') & (
                             data.Destination != '109')
        isinbound = (data.Destination == agent) & (data.Source != '101') & (data.Source != '102') & (
                data.Source != '103') & (
                            data.Source != '104') & (
                            data.Source != '105') & (data.Source != '106') & (data.Source != '107') & (
                            data.Source != '108') & (
                            data.Source != '109')
        date = pd.to_datetime((data["Date"]))

        count = 0
        if status == 'inbound':
            for day in weeklabels:
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & (isinbound)]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & (isinbound)]))
                busy.append(
                    len(data[(data.Status == 'BUSY') & (date.dt.weekday == count) & (isinbound)]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.weekday == count) & (isinbound)]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & (isinbound)]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & (isinbound)]);
                totalbusy = len(data[(data.Status == 'BUSY') & (isinbound)]);
                totalfailed = len(data[(data.Status == 'FAILED') & (isinbound)]);
                count += 1

        elif status == 'outbound':
            for day in weeklabels:
                answeredcall = len(
                    data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isoutbound])
                allcalls = (len(data[(date.dt.weekday == count) & isoutbound]))
                answered.append(
                    len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isoutbound]))
                notanswered.append(
                    len(data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isoutbound]))
                busy.append(len(data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isoutbound]))
                failed.append(
                    len(data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isoutbound]))
                totalanswered = len(data[(data.Status == 'ANSWERED') & isoutbound]);
                totalnoanswer = len(data[(data.Status == 'NO ANSWER') & isoutbound]);
                totalbusy = len(data[(data.Status == 'BUSY') & isoutbound]);
                totalfailed = len(data[(data.Status == 'FAILED') & isoutbound]);
                count += 1

        else:
            for day in weeklabels:
                actualanswered = len(data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & (date.dt.weekday == count) & isinbound])
                answered.append(actualanswered)
                actualnotanswered = len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & (date.dt.weekday == count) & isinbound])
                notanswered.append(actualnotanswered)
                actualbusy = len(
                    data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & (date.dt.weekday == count) & isinbound])
                busy.append(actualbusy)
                actualfailed = len(
                    data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & (date.dt.weekday == count) & isinbound])
                failed.append(actualfailed)

                totalanswered = len(
                    data[(data.Status == 'ANSWERED') & isoutbound]) + len(
                    data[(data.Status == 'ANSWERED') & isinbound])

                totalnoanswer = len(
                    data[(data.Status == 'NO ANSWER') & isoutbound]) + len(
                    data[(data.Status == 'NO ANSWER') & isinbound])
                totalbusy = len(
                    data[(data.Status == 'BUSY') & isoutbound]) + len(
                    data[(data.Status == 'BUSY') & isinbound])
                totalfailed = len(
                    data[(data.Status == 'FAILED') & isoutbound]) + len(
                    data[(data.Status == 'FAILED') & isinbound])
                count += 1

        totalvalues.append(totalanswered)
        totalvalues.append(totalnoanswer)
        totalvalues.append(totalbusy)
        totalvalues.append(totalfailed)
        if totalanswered == 0 & totalnoanswer == 0 & totalbusy == 0 & totalfailed == 0:
            totalvalues.append(0)
        else:
            totalvalues.append(totalanswered / (totalnoanswer + totalbusy + totalfailed))





def databreakdown():
    data = pd.read_csv("cdrmay.csv", low_memory=False, encoding='latin-1')
    ## dd/mm/yyyy format
    currentday = time.strftime("%d/%m/%Y")
    currentweek = time.strftime("%W")
    print(currentweek)

    ## 24 hour format ##
    currenthour = time.strftime("%H:%M:%S")
    currentdate = currentday.split('/')
    currenttime = currenthour.split(':')
    day = int(currentdate[0])
    month = int(currentdate[1])
    year = int(currentdate[2])
    ## Retrieves all Answered on current day
    answeredtoday = data[(data.Call_Status == 'Answered') & (data.Day == day) & (data.Month == month)&(data.Year == year) & (data.Agent != 'Queue')]

    ##Retrieves all not answered on current day
    notansweredtoday = data[(data.Call_Status == 'Not Answered') & (data.Day == day) & (data.Month == month)&(data.Year == year)&(data.Agent != 'Queue')]

    ##Retrieves avg call duration for current day
    dura = data["Duration"][(data.Call_Status == 'Answered') & (data.Day == day) & (data.Month == month)&(data.Year == year) & (data.Agent != 'Queue')]
    duration = dura.tolist()
    if len(duration) != 0:
        avg = sum(duration) / len(duration)
    else:
        avg = 0;

    #Retrieve Previous Day
    date2 = str(year) + '-' + str(month) + '-' + str(day)
    currentdates = datetime.strptime(date2, '%Y-%m-%d') - timedelta(days=1)
    currentdates = datetime.strftime(currentdates, '%Y-%m-%d')
    splitprevday = currentdates.split('-',2)
    prevday = int(splitprevday[2])
    prevmonth = int(splitprevday[1])
    prevyear = int(splitprevday[0])
    ## Retrieves all Answered on previous day
    prevanswered = data[
        (data.Call_Status == 'Answered') & (data.Day == prevday) & (data.Month == prevmonth) & (data.Year == prevyear) & (
                    data.Agent != 'Queue')]

    ##Retrieves all not answered on previous day
    prevnotanswered = data[
        (data.Call_Status == 'Not Answered') & (data.Day == prevday) & (data.Month == prevmonth) & (data.Year == prevyear) & (
                    data.Agent != 'Queue')]

    ##Retrieves avg call duration for previous day
    prevdura = data["Duration"][
        (data.Call_Status == 'Answered') & (data.Day == prevday) & (data.Month == prevmonth) & (data.Year == prevyear) & (
                    data.Agent != 'Queue')]
    prevduration = prevdura.tolist()
    if len(prevduration) != 0:
        prevavg = sum(prevduration) / len(prevduration)
    else:
        prevavg = 0


    #Calculate % Difference in Calls answered
    answereddifference = answeredtoday - prevanswered
    answeredpercentage = (answereddifference / prevanswered) * 100.0

    # Calculate % Difference in Calls Not Answered
    notanswereddifference = notansweredtoday - prevnotanswered
    notansweredpercentage = (notanswereddifference / prevnotanswered) * 100.0


    # Calculate % Difference in Call Durations
    if avg == 0:
        avgpercentage = -100;
    elif prevavg == 0:
        avgpercentage = 100;
    else:
        avgdifference = avg - prevavg
        avgpercentage = (avgdifference / prevavg) * 100.0
avgpercentage = 0
notansweredpercentage=0
answeredpercentage=0
answeredtoday = 0
notansweredtoday = 0
avg = 0
@app.route('/project')
def tester():
    databreakdown()
    return render_template('index.html',totalanswered=answeredtoday,totalmissed=notansweredtoday,avgdura=avg,prevanswered=answeredpercentage
                           ,prevmissed= notansweredpercentage,prevavg=avgpercentage)


@app.route('/weeklyall',methods=['GET', 'POST'])
def weeklyall():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays('all','all',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays(agent,'all', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent=agent)
@app.route('/weeklyinbound',methods=['GET', 'POST'])
def weeklyinbound():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays('all','inbound',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Week(Inbound)', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays(agent,'inbound', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Week(Inbound)', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent=agent)
@app.route('/weeklyoutbound',methods=['GET', 'POST'])
def weeklyoutbound():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays('all','outbound',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Week(Outbound)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatadays(agent,'outbound', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Week(Outbound)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='day',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/monthbreakdownall',methods=['GET', 'POST'])
def breakdownall():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths('all','all',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths(agent,'all', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/monthbreakdowninbound',methods=['GET', 'POST'])
def breakdowninbound():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths('all','all',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths(agent,'all', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/monthbreakdownoutbound',methods=['GET', 'POST'])
def breakdownoutbound():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths('all','all',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatamonths(agent,'all', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call Breakdown by Months', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/allhours',methods=['GET', 'POST'])
def allhours():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatahours('all','all',startdate,enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call time Breakdown(Time Of Day)', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='hour',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent= str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatahours(agent,'all', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Call time Breakdown(Time Of Day)', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='hour',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/inboundhours',methods=['POST','GET'])
def inboundhours():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')
        alldatahours('all','inbound',startdate,enddate);


        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Inbound Call time Breakdown(Time Of Day)', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='hour',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday =str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        if (startday == '0') or (startmonth == '0') or (startyear =='0')  or (endday == '0') or (endmonth == '0') or (endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatahours(agent,'inbound',startdate,enddate);

        a= totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Inbound Call time Breakdown(Time Of Day)', src='inbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered,dayhour='hour',startdate=startdate,enddate=realenddate,agent=agent)

@app.route('/outboundhours',methods=['GET', 'POST'])
def outboundhours():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')
        alldatahours('all','outbound', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Outbound Call time Breakdown(Time Of Day)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered,dayhour='hour',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        alldatahours(agent,'outbound', startdate, enddate);

        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Outbound Call time Breakdown(Time Of Day)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=hourlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='hour', startdate=startdate,
                               enddate=realenddate,agent=agent)

@app.route('/agents',methods=['GET', 'POST'])
def agentlabel():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')
        agentdata('outbound', startdate, enddate);
        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Outbound Call time Breakdown(Time Of Day)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=agentlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered,dayhour='hour',startdate=startdate,enddate=realenddate,agent='All')
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        agentdata('outbound', startdate, enddate);

        a = totalvalues[0]
        n = totalvalues[1]
        b = totalvalues[2]
        f = totalvalues[3]
        return render_template('bar_chart.html', title='Outbound Call time Breakdown(Time Of Day)', src='outbound',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=agentlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='hour', startdate=startdate,
                               enddate=realenddate,agent='all')

@app.route('/123',methods=['GET', 'POST'])
def test():
    if request.method == 'GET':
        date1 = '2018-01-01'
        date2 = '2019-04-01'  # date2 needs a +1

        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        agentsmonths(startdate, enddate);
        a = 1
        n = 2
        b = 3
        f = 4
        return render_template('barchart1.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month', startdate=startdate,
                               enddate=realenddate, agent='All',agentlist=agentslist)
    else:
        result = request.form.to_dict()
        print(result)
        startday = str(result[('startday')])
        startmonth = str(result['startmonth'])
        startyear = str(result['startyear'])
        endday = str(result['endday'])
        endmonth = str(result['endmonth'])
        endyear = str(result['endyear'])
        agent = str(result['agentnumber'])
        print(agent)
        if (startday == '0') or (startmonth == '0') or (startyear == '0') or (endday == '0') or (endmonth == '0') or (
                endyear == '0'):
            date1 = '2018-01-01'
            date2 = '2019-04-01'
        else:
            date1 = startyear + '-' + startmonth + '-' + startday
            print(date1)

            date2 = endyear + '-' + endmonth + '-' + endday
            print(date2)
        startdate = datetime.strptime(date1, '%Y-%m-%d')
        realenddate = datetime.strptime(date2, '%Y-%m-%d')
        realenddate = datetime.strftime(realenddate, '%Y-%m-%d')
        enddate = datetime.strptime(date2, '%Y-%m-%d') + timedelta(days=1)
        startdate = datetime.strftime(startdate, '%Y-%m-%d')
        enddate = datetime.strftime(enddate, '%Y-%m-%d')

        agentsmonths(startdate, enddate);
        a = 1
        n = 2
        b = 3
        f = 4
        return render_template('barchart1.html', title='Call Breakdown by Months', src='all',
                               totalanswered=a, totalfailed=f, totalbusy=b,
                               totalnoanswer=n, labels=monthlabels, answered=answered, busy=busy,
                               failed=failed, notanswered=notanswered, dayhour='month', startdate=startdate,
                               enddate=realenddate, agent=agent,agentlist=agentslist)

@app.route('/alldays',methods=['GET', 'POST'])
def alldays():
    alldatadays('all')

    a = totalvalues[0]
    n = totalvalues[1]
    b = totalvalues[2]
    f = totalvalues[3]
    return render_template('bar_chart.html', title='Call time Breakdown(Day of Week)', src='all',
                           totalanswered=a, totalfailed=f, totalbusy=b,
                           totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                           failed=failed, notanswered=notanswered,dayhour='day')
@app.route('/inbounddays',methods=['GET', 'POST'])
def inbounddays():
    alldatadays('inbound')

    a = totalvalues[0]
    n = totalvalues[1]
    b = totalvalues[2]
    f = totalvalues[3]
    return render_template('bar_chart.html', title='Inbound Call time Breakdown(Day of Week)', src='inbound',
                           totalanswered=a, totalfailed=f, totalbusy=b,
                           totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                           failed=failed, notanswered=notanswered,dayhour='day')
@app.route('/outbounddays',methods=['GET', 'POST'])
def outbounddays():
    alldatadays('outbound')

    a = totalvalues[0]
    n = totalvalues[1]
    b = totalvalues[2]
    f = totalvalues[3]
    return render_template('bar_chart.html', title='Outbound Call time Breakdown(Day of Week)', src='outbound',
                           totalanswered=a, totalfailed=f, totalbusy=b,
                           totalnoanswer=n, labels=weeklabels, answered=answered, busy=busy,
                           failed=failed, notanswered=notanswered,dayhour='day')
@app.route('/boxplotall')
def boxplotall():
    alldataboxplot('all')
    print(biglist[0])
    return render_template('boxplot.html',data=biglist,title='Box Plot All Answered Calls')
@app.route('/boxplotinbound')
def boxplotinbound():
    alldataboxplot('inbound')
    print(biglist[0])
    return render_template('boxplot.html',data=biglist,title='Box Plot Inbound Answered Calls')
@app.route('/boxplotoutbound')
def boxplotoutbound():
    alldataboxplot('outbound')
    print(biglist[0])
    return render_template('boxplot.html',data=biglist,title='Box Plot Outbound Answered Calls')

@app.route('/')
def landing_page():
    return render_template('index.html')


@app.route('/testtest')
def testing():
    writestatus()
    return render_template('index.html')


if __name__ == '__main__':
    app.run()



